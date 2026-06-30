"""엑셀 통합문서 구조 분석 엔진.

openpyxl 로 .xlsx 를 열어 시트별 수식을 추출하고, 수식 안의 참조를 모아
 1) 시트 간 의존 관계(시트 단위)
 2) 셀 간 의존 관계(셀 단위 — 상류/하류 추적용)
를 함께 만든다.

화살표 방향(데이터 흐름):
    셀이 SheetB!A1 을 참조하면 데이터는 SheetB -> (현재 시트) 로 흐른다.
    따라서 의존 엣지는 (참조된 셀/시트=출처) -> (참조하는 셀/시트=소비) 방향으로 저장한다.
    "입력 -> 가공 -> 산출" 흐름이 자연스럽게 보인다.
"""

from __future__ import annotations

import os
from collections import defaultdict
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

from .formula_parser import extract_cell_references


# 범위 참조를 하류(dependents) 전개할 때의 셀 수 상한.
# 이보다 큰 범위(예: SUM(A1:A100000))는 전개하지 않아 폭발을 막는다(범위 단위로만 유지).
MAX_RANGE_CELLS = 4096


@dataclass
class SheetInfo:
    """한 시트의 요약 정보."""

    name: str
    index: int
    state: str                    # visible / hidden / veryHidden
    dimensions: str = ""          # 사용 범위 (예: "A1:AM81")
    n_rows: int = 0               # 사용 범위의 행 수 (span)
    n_cols: int = 0               # 사용 범위의 열 수 (span)
    max_row: int = 0              # 절대 최대 행 (A1 기준)
    max_col: int = 0              # 절대 최대 열 (A1 기준)
    formula_count: int = 0        # 수식 셀 수
    number_count: int = 0         # 숫자/날짜 셀 수
    text_count: int = 0           # 텍스트 셀 수
    # 이 시트의 모든 수식: (셀주소, 수식원문)
    formulas: list[tuple[str, str]] = field(default_factory=list)
    # 이 시트가 참조하는 다른 시트 이름들
    references: set[str] = field(default_factory=set)

    @property
    def grid_area(self) -> int:
        return self.max_row * self.max_col

    @property
    def formula_ratio(self) -> float:
        """수식 비율 = 수식 셀 수 / (절대 최대행 × 절대 최대열)."""
        area = self.grid_area
        return (self.formula_count / area) if area else 0.0


@dataclass
class Dependency:
    """from_sheet -> to_sheet 데이터 흐름 (to_sheet 가 from_sheet 를 참조)."""

    from_sheet: str               # 데이터 출처 (참조 대상 시트)
    to_sheet: str                 # 데이터 소비 (수식이 있는 시트)
    links: list[tuple[str, str]] = field(default_factory=list)

    @property
    def weight(self) -> int:
        return len(self.links)


@dataclass
class CellRef:
    """셀 단위 선행참조 한 건(precedent). 외부참조면 sheet="" , is_external=True."""

    sheet: str
    ref: str
    is_range: bool = False
    is_external: bool = False
    raw: str = ""


@dataclass
class CellNode:
    """셀 단위 그래프의 노드 하나. key = f'{sheet}!{coord}'."""

    sheet: str
    coord: str
    formula: str | None = None
    value: str | None = None
    precedents: list[CellRef] = field(default_factory=list)   # 이 셀이 참조하는 것들(상류)
    dependents: list[str] = field(default_factory=list)       # 이 셀을 참조하는 셀 key(하류)


@dataclass
class WorkbookAnalysis:
    """통합문서 전체 분석 결과."""

    file_path: str
    sheets: list[SheetInfo] = field(default_factory=list)
    dependencies: list[Dependency] = field(default_factory=list)
    # 셀 단위 그래프: key("시트!셀") -> CellNode
    cells: dict[str, CellNode] = field(default_factory=dict)
    # 시트쌍 드릴다운: "from->to" -> [{dst, src, formula}]
    edge_cells: dict[str, list[dict]] = field(default_factory=dict)
    external_refs: int = 0        # 외부 통합문서 참조 개수
    self_refs: int = 0            # 자기 시트 내 참조 개수
    range_capped: int = 0         # 상한 초과로 하류 전개를 생략한 범위 참조 수
    warnings: list[str] = field(default_factory=list)

    @property
    def file_name(self) -> str:
        return os.path.basename(self.file_path)

    @property
    def total_formulas(self) -> int:
        return sum(s.formula_count for s in self.sheets)

    def input_sheets(self) -> list[str]:
        """다른 시트를 참조하지 않는(=의존 대상이 없는) 시트 = 입력단 후보."""
        consumers = {d.to_sheet for d in self.dependencies}
        return [s.name for s in self.sheets if s.name not in consumers]

    def output_sheets(self) -> list[str]:
        """다른 시트가 참조하지 않는 시트 = 최종 산출 후보."""
        sources = {d.from_sheet for d in self.dependencies}
        return [s.name for s in self.sheets if s.name not in sources]

    def find_cycle(self) -> list[str]:
        """시트 간 순환 참조를 하나 찾아 경로로 반환한다. 없으면 빈 리스트."""
        adj: dict[str, list[str]] = defaultdict(list)
        for d in self.dependencies:
            adj[d.from_sheet].append(d.to_sheet)

        WHITE, GRAY, BLACK = 0, 1, 2
        color = {s.name: WHITE for s in self.sheets}
        stack: list[str] = []

        def dfs(u: str) -> list[str]:
            color[u] = GRAY
            stack.append(u)
            for v in adj.get(u, []):
                if color.get(v) == GRAY:
                    return stack[stack.index(v):] + [v]
                if color.get(v) == WHITE:
                    found = dfs(v)
                    if found:
                        return found
            color[u] = BLACK
            stack.pop()
            return []

        for s in self.sheets:
            if color[s.name] == WHITE:
                found = dfs(s.name)
                if found:
                    return found
        return []

    def has_cycle(self) -> bool:
        return bool(self.find_cycle())


def _formula_text(value) -> str | None:
    """셀 값에서 수식 원문을 꺼낸다. 수식이 아니면 None."""
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _range_cells(ref: str, cap: int) -> list[str] | None:
    """범위 문자열을 멤버 셀 좌표 리스트로 전개. 너무 크거나 전체열/행이면 None."""
    try:
        c1, r1, c2, r2 = range_boundaries(ref)
    except Exception:
        return None
    if None in (c1, r1, c2, r2):          # A:A, 1:1 같은 전체열/행 → 전개 생략
        return None
    if (r2 - r1 + 1) * (c2 - c1 + 1) > cap:
        return None
    return [f"{get_column_letter(c)}{r}" for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)]


def analyze_workbook(file_path: str) -> WorkbookAnalysis:
    """.xlsx 파일을 분석해 WorkbookAnalysis 를 반환한다.

    .xls(구버전)는 openpyxl 로 열 수 없으므로 명확한 안내 메시지와 함께
    예외를 던진다.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        raise ValueError(
            "구버전 .xls 파일은 직접 분석할 수 없습니다.\n"
            "엑셀에서 파일을 열고 [다른 이름으로 저장] → 형식을 "
            "'Excel 통합 문서(*.xlsx)'로 선택해 저장한 뒤, 그 .xlsx 파일을 "
            "다시 분석해 주세요."
        )
    if ext not in (".xlsx", ".xlsm"):
        raise ValueError(
            f"지원하지 않는 형식입니다: {ext or '(확장자 없음)'}\n"
            ".xlsx 또는 .xlsm 파일을 사용해 주세요."
        )

    analysis = WorkbookAnalysis(file_path=file_path)

    # data_only=False: 계산값이 아니라 수식 원문을 읽는다.
    wb = load_workbook(filename=file_path, data_only=False, read_only=False)
    known_sheets = wb.sheetnames

    edge_links: dict[tuple[str, str], list[tuple[str, str]]] = defaultdict(list)
    cells = analysis.cells
    # 수식 셀 목록을 모아 2차 패스에서 참조를 푼다: (sheet, coord, formula)
    formula_cells: list[tuple[str, str, str]] = []

    def node(sheet: str, coord: str) -> CellNode:
        key = f"{sheet}!{coord}"
        n = cells.get(key)
        if n is None:
            n = CellNode(sheet=sheet, coord=coord)
            cells[key] = n
        return n

    # ── 1차 패스: 시트/셀 통계 + 노드(수식·값) 기록 ──────────────────────────
    for idx, ws in enumerate(wb.worksheets):
        min_row, max_row = ws.min_row or 1, ws.max_row or 1
        min_col, max_col = ws.min_column or 1, ws.max_column or 1
        info = SheetInfo(
            name=ws.title,
            index=idx,
            state=ws.sheet_state,
            dimensions=ws.dimensions,
            n_rows=max_row - min_row + 1,
            n_cols=max_col - min_col + 1,
            max_row=max_row,
            max_col=max_col,
        )

        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                formula = _formula_text(value)
                if formula is not None:
                    info.formula_count += 1
                    info.formulas.append((cell.coordinate, formula))
                    node(ws.title, cell.coordinate).formula = formula
                    formula_cells.append((ws.title, cell.coordinate, formula))
                elif isinstance(value, str):
                    info.text_count += 1
                    node(ws.title, cell.coordinate).value = value
                else:
                    info.number_count += 1
                    node(ws.title, cell.coordinate).value = str(value)

        analysis.sheets.append(info)

    name_to_info = {s.name: s for s in analysis.sheets}

    # ── 2차 패스: 수식의 참조를 풀어 셀/시트 그래프 구성 ──────────────────────
    for sheet, coord, formula in formula_cells:
        fkey = f"{sheet}!{coord}"
        fnode = cells[fkey]
        seen_prec = set()
        for r in extract_cell_references(formula, sheet, known_sheets):
            if r.is_external:
                analysis.external_refs += 1
                fnode.precedents.append(CellRef("", r.ref, r.is_range, True, r.raw))
                continue

            skey = f"{r.sheet}!{r.ref}"
            if skey == fkey:                      # 자기 셀 자기참조는 생략
                continue
            if skey not in seen_prec:
                seen_prec.add(skey)
                fnode.precedents.append(CellRef(r.sheet, r.ref, r.is_range))

            if r.sheet == sheet:
                analysis.self_refs += 1
            else:
                name_to_info[sheet].references.add(r.sheet)
                edge_links[(r.sheet, sheet)].append((coord, formula))
                analysis.edge_cells.setdefault(f"{r.sheet}->{sheet}", []).append(
                    {"dst": fkey, "src": skey, "formula": formula}
                )

            # 하류(dependents): 참조된 셀(범위면 멤버들)이 이 수식셀을 dependent 로 가짐
            members = _range_cells(r.ref, MAX_RANGE_CELLS)
            if members is None:
                if r.is_range:
                    analysis.range_capped += 1
                members = [r.ref] if not r.is_range else []
            for m in members:
                node(r.sheet, m).dependents.append(fkey)

    # dependents 중복 제거
    for n in cells.values():
        if n.dependents:
            n.dependents = sorted(set(n.dependents))

    wb.close()

    for (frm, to), links in edge_links.items():
        analysis.dependencies.append(
            Dependency(from_sheet=frm, to_sheet=to, links=links)
        )
    analysis.dependencies.sort(key=lambda d: (-d.weight, d.from_sheet, d.to_sheet))

    if not analysis.dependencies:
        analysis.warnings.append(
            "시트 간 참조 수식을 찾지 못했습니다. 시트들이 서로 독립적이거나, "
            "참조가 외부 파일/이름 정의를 통한 것일 수 있습니다."
        )
    if analysis.range_capped:
        analysis.warnings.append(
            f"매우 큰 범위 참조 {analysis.range_capped}개는 성능을 위해 셀 단위 "
            "하류 추적에서 범위 단위로만 처리했습니다."
        )

    return analysis
