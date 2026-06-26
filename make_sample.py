"""검증용 샘플 통합문서 생성.

시나리오: 입력(A) -> 가공(B) -> 표(C, D). D 는 단가표(F)도 참조. 최종 요약(E)는 C, D 참조.
한글 시트명/공백 시트명/3D 참조/외부참조도 섞어서 파서를 시험한다.
"""

from openpyxl import Workbook

wb = Workbook()

a = wb.active
a.title = "입력"
a["A1"] = "수량"
a["A2"] = 10
a["A3"] = 20

b = wb.create_sheet("가공 단계")          # 공백 포함 시트명
b["A1"] = "=입력!A2*2"
b["A2"] = "='입력'!A3*2"
b["B1"] = "=SUM(입력!A2:A3)"

c = wb.create_sheet("표C")
c["A1"] = "='가공 단계'!A1"
c["A2"] = "='가공 단계'!B1"

f = wb.create_sheet("단가표")
f["A1"] = 1500

d = wb.create_sheet("표D")
d["A1"] = "='가공 단계'!A1*단가표!A1"     # 두 시트 동시 참조
d["A2"] = "=표C!A1+단가표!A1"

e = wb.create_sheet("요약")
e["A1"] = "=표C!A1+표D!A1"
e["A2"] = "=SUM('가공 단계:표D'!A1)"       # 3D 참조
e["A3"] = "=[1]외부!A1"                     # 외부 통합문서 참조
e["A4"] = "=요약!A1"                        # 자기 참조

wb.save("sample.xlsx")
print("sample.xlsx 생성 완료")
